import streamlit as st
import pandas as pd
import re
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# ─── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CAN Bus Analyzer",
    page_icon="🚌",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── THEME CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Dark engineering feel */
:root {
    --bg-deep: #0d1117;
    --bg-card: #161b22;
    --bg-panel: #1c2128;
    --accent: #58a6ff;
    --accent2: #3fb950;
    --accent3: #f78166;
    --accent4: #d2a8ff;
    --text: #e6edf3;
    --text-dim: #8b949e;
    --border: #30363d;
    --tx-color: #3fb950;
    --rx-color: #58a6ff;
    --ext-color: #d2a8ff;
    --rem-color: #e3b341;
}

/* Hide default streamlit chrome */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* Main background */
.stApp { background: var(--bg-deep); }
section[data-testid="stSidebar"] { background: var(--bg-card) !important; border-right: 1px solid var(--border); }

/* Sidebar header */
.sidebar-logo {
    font-size: 1.3rem; font-weight: 700; color: var(--accent);
    padding: 0.5rem 0 1rem 0; letter-spacing: 0.05em;
    border-bottom: 1px solid var(--border); margin-bottom: 1rem;
}

/* KPI cards */
.kpi-row { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 1rem; }
.kpi-card {
    flex: 1; min-width: 130px;
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 12px 16px;
    text-align: center;
}
.kpi-card.tx { border-top: 3px solid var(--tx-color); }
.kpi-card.rx { border-top: 3px solid var(--rx-color); }
.kpi-card.ext { border-top: 3px solid var(--ext-color); }
.kpi-card.total { border-top: 3px solid var(--accent); }
.kpi-card.ids { border-top: 3px solid var(--accent4); }
.kpi-card.dur { border-top: 3px solid var(--rem-color); }

.kpi-val { font-size: 1.7rem; font-weight: 700; color: var(--text); line-height: 1.1; }
.kpi-label { font-size: 0.72rem; color: var(--text-dim); margin-top: 3px; text-transform: uppercase; letter-spacing: 0.05em; }

/* Section headers */
.sec-header {
    font-size: 0.8rem; font-weight: 600; color: var(--text-dim);
    text-transform: uppercase; letter-spacing: 0.08em;
    padding: 0.5rem 0 0.3rem 0;
    border-bottom: 1px solid var(--border);
    margin: 0.5rem 0 0.8rem 0;
}

/* ID badge pills */
.id-pill {
    display: inline-block;
    background: #1f2937; border: 1px solid var(--border);
    border-radius: 4px; padding: 1px 6px;
    font-family: monospace; font-size: 0.8rem; color: var(--accent);
    font-weight: 600;
}

/* Direction badges */
.badge-tx { background:#0d2818; color:#3fb950; border:1px solid #3fb950; border-radius:3px; padding:1px 6px; font-size:0.75rem; font-weight:700; }
.badge-rx { background:#0c1e35; color:#58a6ff; border:1px solid #58a6ff; border-radius:3px; padding:1px 6px; font-size:0.75rem; font-weight:700; }

/* Frame type badges */
.badge-std { background:#1a1a2e; color:#c3c3e0; border:1px solid #444; border-radius:3px; padding:1px 5px; font-size:0.7rem; }
.badge-ext { background:#1e1030; color:#d2a8ff; border:1px solid #7c4dff; border-radius:3px; padding:1px 5px; font-size:0.7rem; }
.badge-rem { background:#1e1500; color:#e3b341; border:1px solid #e3b341; border-radius:3px; padding:1px 5px; font-size:0.7rem; }

/* Info box */
.info-box {
    background: var(--bg-panel); border: 1px solid var(--border);
    border-left: 3px solid var(--accent);
    border-radius: 6px; padding: 0.8rem 1rem;
    font-size: 0.85rem; color: var(--text-dim); margin-bottom: 1rem;
}

/* Byte grid */
.byte-grid {
    display: flex; gap: 4px; flex-wrap: wrap; font-family: monospace;
}
.byte-cell {
    width: 34px; height: 28px; display: flex; align-items: center;
    justify-content: center; border-radius: 4px;
    font-size: 0.75rem; font-weight: 600;
}
.byte-zero { background: #1c2128; color: #444d56; border: 1px solid #30363d; }
.byte-nonzero { background: #132b1e; color: #3fb950; border: 1px solid #3fb950; }

/* Metric deltas override */
[data-testid="stMetricDelta"] { font-size: 0.7rem; }

/* Tab styling */
.stTabs [data-baseweb="tab-list"] {
    background: var(--bg-card);
    border-bottom: 1px solid var(--border);
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    background: transparent; color: var(--text-dim);
    border: none; font-size: 0.82rem; font-weight: 500;
}
.stTabs [aria-selected="true"] {
    background: var(--bg-panel) !important; color: var(--accent) !important;
    border-bottom: 2px solid var(--accent) !important;
}

/* Upload zone */
[data-testid="stFileUploader"] {
    background: var(--bg-card);
    border: 2px dashed var(--border);
    border-radius: 8px;
}

/* Dataframe */
[data-testid="stDataFrame"] { border-radius: 8px; }

/* Sidebar filter labels */
.filter-label {
    font-size: 0.72rem; font-weight: 600; color: var(--text-dim);
    text-transform: uppercase; letter-spacing: 0.06em;
    margin-bottom: 2px;
}

/* Session info table */
.meta-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
.meta-table td { padding: 5px 8px; border-bottom: 1px solid var(--border); color: var(--text); }
.meta-table td:first-child { color: var(--text-dim); font-size: 0.75rem; text-transform: uppercase; letter-spacing:0.04em; width: 40%; }

/* Timing table */
.timing-row { display: flex; align-items: center; gap: 8px; margin-bottom: 6px; padding: 6px 10px; background: var(--bg-panel); border-radius: 6px; border: 1px solid var(--border); }
.timing-id { font-family: monospace; font-weight: 700; color: var(--accent); min-width: 110px; font-size: 0.85rem; }
.timing-bar-wrap { flex: 1; background: #21262d; border-radius: 3px; height: 8px; overflow: hidden; }
.timing-bar { height: 100%; border-radius: 3px; background: linear-gradient(90deg, #58a6ff, #3fb950); }
.timing-val { font-size: 0.78rem; color: var(--text-dim); min-width: 80px; text-align: right; }

/* Warning banner */
.warn-banner {
    background: #2d1e00; border: 1px solid #e3b341; border-left: 3px solid #e3b341;
    border-radius: 6px; padding: 8px 12px; font-size: 0.82rem; color: #e3b341; margin-bottom: 8px;
}

/* Download button override */
.stDownloadButton > button {
    background: linear-gradient(90deg, #1a5276, #1a3a5c) !important;
    color: #58a6ff !important;
    border: 1px solid #58a6ff !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    width: 100% !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(90deg, #1f6391, #1a4a7c) !important;
}
</style>
""", unsafe_allow_html=True)


# ─── PARSER ────────────────────────────────────────────────────────────────────
def parse_can_log(content: str):
    lines = content.splitlines()
    meta, rows = {}, []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith("***"):
            m = re.match(r"\*\*\*START DATE AND TIME (.+?)\*\*\*", line)
            if m: meta["Date / Time"] = m.group(1).strip()
            m = re.match(r"\*\*\*PROTOCOL (.+?)\*\*\*", line)
            if m: meta["Protocol"] = m.group(1).strip()
            m = re.match(r"\*\*\*BUSMASTER Ver (.+?)\*\*\*", line)
            if m: meta["Logger"] = "BUSMASTER v" + m.group(1).strip()
            if "500000 bps" in line or "bps" in line:
                m = re.search(r"(\d+ bps)", line)
                if m: meta["Baud Rate"] = m.group(1)
            m = re.search(r"Kvaser[^,]+", line)
            if m: meta["Interface"] = m.group(0).strip()
            m = re.search(r"Serial Number- (\S+)", line)
            if m: meta["Serial No."] = m.group(1)
            if "SYSTEM MODE" in line: meta["Mode"] = "System Mode"
            if "HEX" in line and "DATABASE" not in line: meta["Display Format"] = "HEX"
            continue

        parts = line.split()
        if len(parts) < 6:
            continue
        ts_str = parts[0]
        direction = parts[1]
        channel   = parts[2]
        can_id    = parts[3]
        ft_raw    = parts[4]
        try:
            dlc = int(parts[5])
        except ValueError:
            dlc = 0
        data_bytes = parts[6:]

        m = re.match(r"(\d+):(\d+):(\d+):(\d+)", ts_str)
        if m:
            h, mi, s, ms = map(int, m.groups())
            ts_ms = h*3600000 + mi*60000 + s*1000 + ms
        else:
            ts_ms = None

        if "sr" in ft_raw:
            frame_type = "Remote"
        elif "x" in ft_raw:
            frame_type = "Extended"
        else:
            frame_type = "Standard"

        byte_dict = {f"B{i}": (data_bytes[i] if i < len(data_bytes) else "--") for i in range(8)}
        data_hex = " ".join(data_bytes) if data_bytes else "-- -- -- -- -- -- -- --"

        rows.append({
            "Timestamp":     ts_str,
            "ts_ms":         ts_ms,
            "Direction":     direction,
            "Channel":       channel,
            "CAN ID":        can_id,
            "Frame":         frame_type,
            "DLC":           dlc,
            "Data (Hex)":    data_hex,
            **byte_dict,
        })

    df = pd.DataFrame(rows)
    if not df.empty and df["ts_ms"].notna().any():
        t0 = df["ts_ms"].min()
        df["Rel Time (ms)"] = (df["ts_ms"] - t0).round(1)
    else:
        df["Rel Time (ms)"] = None
    return meta, df


# ─── TIMING ANALYSIS ───────────────────────────────────────────────────────────
def timing_analysis(df: pd.DataFrame):
    results = []
    for cid, grp in df.groupby("CAN ID"):
        grp = grp.sort_values("ts_ms").copy()
        diffs = grp["ts_ms"].diff().dropna()
        if len(diffs) < 2:
            continue
        results.append({
            "CAN ID":        cid,
            "Direction":     grp["Direction"].mode()[0],
            "Frame":         grp["Frame"].mode()[0],
            "Count":         len(grp),
            "Mean (ms)":     round(diffs.mean(), 2),
            "Min (ms)":      round(diffs.min(), 2),
            "Max (ms)":      round(diffs.max(), 2),
            "Std Dev (ms)":  round(diffs.std(), 2),
            "Est. Period":   f"~{round(diffs.mean())} ms",
            "Est. Freq":     f"{round(1000/diffs.mean(), 1)} Hz" if diffs.mean() > 0 else "-",
            "Jitter (ms)":   round(diffs.std(), 2),
        })
    return pd.DataFrame(results)


# ─── EXCEL EXPORT ──────────────────────────────────────────────────────────────
def build_excel(meta, df):
    wb = Workbook()

    C_DARK="0D1117"; C_CARD="161B22"; C_HDR="1C2128"
    C_ACCENT="58A6FF"; C_TX="3FB950"; C_RX="58A6FF"
    C_TEXT="E6EDF3"; C_DIM="8B949E"; C_BORDER="30363D"
    C_ROW1="161B22"; C_ROW2="1C2128"

    thin=Side(style="thin",color=C_BORDER)
    brd=Border(left=thin,right=thin,top=thin,bottom=thin)

    def hf(sz=10,bold=True,color=C_TEXT): return Font(name="Consolas",size=sz,bold=bold,color=color)
    def bf(sz=9,bold=False,color=C_TEXT): return Font(name="Consolas",size=sz,bold=bold,color=color)
    def fl(c): return PatternFill("solid",fgColor=c)
    def ac(): return Alignment(horizontal="center",vertical="center")
    def al(): return Alignment(horizontal="left",vertical="center")
    def ar(): return Alignment(horizontal="right",vertical="center")

    def make_title(ws, text, ncols, bg=C_DARK):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c=ws["A1"]; c.value=text
        c.font=Font(name="Consolas",size=13,bold=True,color=C_ACCENT)
        c.fill=fl(bg); c.alignment=al(); ws.row_dimensions[1].height=28

    def make_header(ws, cols, row=2):
        for ci,(name,w) in enumerate(cols,1):
            c=ws.cell(row=row,column=ci,value=name)
            c.font=hf(9); c.fill=fl(C_HDR); c.alignment=ac(); c.border=brd
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.row_dimensions[row].height=20

    def write_df_rows(ws, dataframe, col_defs, start_row=3, tx_green=False):
        """Write a DataFrame to a worksheet safely using iloc."""
        col_names = [c[0] for c in col_defs]
        df_cols = list(dataframe.columns)
        for ri, (_, row) in enumerate(dataframe.iterrows(), start=start_row):
            d = row.get("Direction","")
            if tx_green:
                bg = ("0D2818" if ri%2==0 else "132b1e") if d=="Tx" else ("0C1E35" if ri%2==0 else "0f1e33")
            else:
                bg = C_ROW2 if ri%2==0 else C_ROW1
            txt_color = C_TX if d=="Tx" else (C_RX if d=="Rx" else C_TEXT)
            for ci, cname in enumerate(col_names, 1):
                # Map display col name back to dataframe col
                mapped = {"Dir":"Direction","Ch":"Channel","B0":"B0","B1":"B1","B2":"B2",
                          "B3":"B3","B4":"B4","B5":"B5","B6":"B6","B7":"B7"}.get(cname, cname)
                val = row.get(mapped, "") if mapped in df_cols else ""
                if val is None: val = ""
                cell = ws.cell(ri, ci, val)
                cell.font = bf(9, color=txt_color if cname in ("CAN ID","Dir","Direction") else C_TEXT)
                cell.fill = fl(bg)
                cell.alignment = al() if cname in ("Timestamp","Data (Hex)","Rel Time (ms)") else ac()
                cell.border = brd

    # ── Sheet 1: Summary ──
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    make_title(ws, "  CAN BUS LOG — ANALYSIS REPORT", 2)

    r=3
    ws.merge_cells(f"A{r}:B{r}")
    lbl=ws.cell(r,1,"SESSION INFO"); lbl.font=hf(8,color=C_DIM); lbl.fill=fl(C_HDR); lbl.alignment=al(); r+=1
    for k,v in meta.items():
        ws.cell(r,1,k).font=bf(color=C_DIM); ws.cell(r,1).fill=fl(C_ROW2); ws.cell(r,1).border=brd; ws.cell(r,1).alignment=al()
        ws.cell(r,2,str(v)).font=bf(bold=True); ws.cell(r,2).fill=fl(C_ROW1); ws.cell(r,2).border=brd; ws.cell(r,2).alignment=al()
        r+=1

    r+=1
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r,1,"LOG STATISTICS").font=hf(8,color=C_DIM); ws.cell(r,1).fill=fl(C_HDR); ws.cell(r,1).alignment=al(); r+=1

    dur_ms = df["ts_ms"].max()-df["ts_ms"].min() if df["ts_ms"].notna().any() else 0
    dur_s  = dur_ms/1000
    stats=[
        ("Total Messages",       len(df)),
        ("Transmitted (Tx)",     len(df[df["Direction"]=="Tx"])),
        ("Received (Rx)",        len(df[df["Direction"]=="Rx"])),
        ("Unique CAN IDs",       df["CAN ID"].nunique()),
        ("Standard Frames",      len(df[df["Frame"]=="Standard"])),
        ("Extended Frames",      len(df[df["Frame"]=="Extended"])),
        ("Remote Frames",        len(df[df["Frame"]=="Remote"])),
        ("Log Duration (s)",     round(dur_s,3)),
        ("Avg Msg Rate (msg/s)", round(len(df)/dur_s,1) if dur_s>0 else "-"),
        ("Most Active CAN ID",   df["CAN ID"].value_counts().index[0]),
    ]
    for k,v in stats:
        ws.cell(r,1,k).font=bf(color=C_DIM); ws.cell(r,1).fill=fl(C_ROW2); ws.cell(r,1).border=brd; ws.cell(r,1).alignment=al()
        ws.cell(r,2,str(v)).font=bf(bold=True,color=C_ACCENT); ws.cell(r,2).fill=fl(C_ROW1); ws.cell(r,2).border=brd; ws.cell(r,2).alignment=ar()
        r+=1

    # ── Sheet 2: All Messages ──
    ws2 = wb.create_sheet("All Messages")
    ws2.sheet_view.showGridLines=False; ws2.freeze_panes="A3"
    msg_cols=[("Timestamp",18),("Rel Time (ms)",14),("Dir",6),("Ch",5),("CAN ID",13),
              ("Frame",11),("DLC",5),("Data (Hex)",26),
              ("B0",7),("B1",7),("B2",7),("B3",7),("B4",7),("B5",7),("B6",7),("B7",7)]
    make_title(ws2,"  ALL CAN MESSAGES — "+str(len(df))+" ROWS",len(msg_cols))
    make_header(ws2, msg_cols)
    ws2.auto_filter.ref=f"A2:{get_column_letter(len(msg_cols))}2"

    # Build display df with renamed cols
    exp_df = df.rename(columns={"Direction":"Dir","Channel":"Ch"})[
        ["Timestamp","Rel Time (ms)","Dir","Ch","CAN ID","Frame","DLC","Data (Hex)",
         "B0","B1","B2","B3","B4","B5","B6","B7"]
    ].copy()
    exp_df["Dir"] = df["Direction"]  # keep for color logic
    # add Direction back for color
    exp_df2 = exp_df.copy()
    exp_df2["Direction"] = df["Direction"].values

    for ri,(_, row) in enumerate(exp_df2.iterrows(), start=3):
        d = row.get("Direction","")
        bg = ("0D2818" if ri%2==0 else "132b1e") if d=="Tx" else ("0C1E35" if ri%2==0 else "0f1e33")
        cols_in_sheet = [c[0] for c in msg_cols]
        for ci, cname in enumerate(cols_in_sheet, 1):
            val = row.get(cname, "")
            if val is None: val = ""
            cell = ws2.cell(ri, ci, val)
            cell.font = bf(9, color=C_TX if d=="Tx" else C_TEXT)
            cell.fill = fl(bg)
            cell.alignment = al() if cname in ("Timestamp","Data (Hex)","Rel Time (ms)") else ac()
            cell.border = brd

    # ── Sheet 3: Timing Analysis ──
    ws3=wb.create_sheet("Timing Analysis")
    ws3.sheet_view.showGridLines=False; ws3.freeze_panes="A3"
    t_cols=[("CAN ID",14),("Direction",10),("Frame",12),("Count",8),
            ("Mean (ms)",11),("Min (ms)",10),("Max (ms)",10),("Std Dev (ms)",12),
            ("Est. Period",13),("Est. Freq",11),("Jitter (ms)",11)]
    make_title(ws3,"  TIMING ANALYSIS — PER CAN ID",len(t_cols))
    make_header(ws3, t_cols)
    ta=timing_analysis(df)

    if not ta.empty:
        mean_max=ta["Mean (ms)"].max(); jit_max=ta["Jitter (ms)"].max()
        for ri,(_, row) in enumerate(ta.iterrows(), start=3):
            bg=C_ROW2 if ri%2==0 else C_ROW1
            for ci,cname in enumerate([c[0] for c in t_cols],1):
                val=row.get(cname,"")
                if val is None: val=""
                cell=ws3.cell(ri,ci,val)
                # Manual color scale — no matplotlib
                if cname=="Mean (ms)" and isinstance(val,(int,float)) and mean_max>0:
                    ratio=min(val/mean_max,1.0)
                    r_=int(30+ratio*200); g_=int(180-ratio*130)
                    cell.fill=PatternFill("solid",fgColor=f"{r_:02X}{g_:02X}32")
                elif cname=="Jitter (ms)" and isinstance(val,(int,float)) and jit_max>0:
                    ratio=min(val/jit_max,1.0)
                    r_=int(30+ratio*220); g_=int(60-ratio*40)
                    cell.fill=PatternFill("solid",fgColor=f"{r_:02X}{g_:02X}32")
                else:
                    cell.fill=fl(bg)
                cell.font=bf(9,color=C_ACCENT if cname=="CAN ID" else C_TEXT)
                cell.alignment=ac(); cell.border=brd

    # ── Sheet 4: ID Breakdown ──
    ws4=wb.create_sheet("ID Breakdown")
    ws4.sheet_view.showGridLines=False; ws4.freeze_panes="A3"
    b_cols=[("CAN ID",14),("Total",8),("Tx",7),("Rx",7),("Frame Type",14),
            ("% Traffic",11),("Unique Payloads",16),("First Seen",18),("Last Seen",18)]
    make_title(ws4,"  CAN ID TRAFFIC BREAKDOWN",len(b_cols))
    make_header(ws4, b_cols)
    total=len(df)
    for ri,(cid,grp) in enumerate(df.groupby("CAN ID"),start=3):
        bg=C_ROW2 if ri%2==0 else C_ROW1
        pct=round(len(grp)/total*100,1)
        vals=[cid,len(grp),len(grp[grp["Direction"]=="Tx"]),len(grp[grp["Direction"]=="Rx"]),
              grp["Frame"].mode()[0] if not grp.empty else "-",f"{pct}%",
              grp["Data (Hex)"].nunique(),grp["Timestamp"].iloc[0],grp["Timestamp"].iloc[-1]]
        for ci,v in enumerate(vals,1):
            cell=ws4.cell(ri,ci,v)
            cell.font=bf(9,color=C_ACCENT if ci==1 else C_TEXT)
            cell.fill=fl(bg); cell.alignment=ac(); cell.border=brd
    ws4.auto_filter.ref=f"A2:{get_column_letter(len(b_cols))}{len(df.groupby('CAN ID'))+2}"

    # ── Sheet 5: Tx Messages ──
    ws5=wb.create_sheet("Tx Messages")
    ws5.sheet_view.showGridLines=False; ws5.freeze_panes="A3"
    make_title(ws5,"  TRANSMITTED (Tx) MESSAGES",len(msg_cols),"0D2818")
    make_header(ws5, msg_cols)
    ws5.auto_filter.ref=f"A2:{get_column_letter(len(msg_cols))}2"
    tx_df=df[df["Direction"]=="Tx"].reset_index(drop=True)
    for ri,(_, row) in enumerate(tx_df.iterrows(), start=3):
        bg="0D2818" if ri%2==0 else "132b1e"
        cols_in_sheet=[c[0] for c in msg_cols]
        for ci,cname in enumerate(cols_in_sheet,1):
            mapped={"Dir":"Direction","Ch":"Channel"}.get(cname,cname)
            val=row.get(mapped,"")
            if val is None: val=""
            if cname=="Dir": val="Tx"
            cell=ws5.cell(ri,ci,val)
            cell.font=bf(9,color=C_TX); cell.fill=fl(bg)
            cell.alignment=al() if cname in ("Timestamp","Data (Hex)","Rel Time (ms)") else ac()
            cell.border=brd

    # ── Sheet 6: Rx Messages ──
    ws6=wb.create_sheet("Rx Messages")
    ws6.sheet_view.showGridLines=False; ws6.freeze_panes="A3"
    make_title(ws6,"  RECEIVED (Rx) MESSAGES",len(msg_cols),"0C1E35")
    make_header(ws6, msg_cols)
    ws6.auto_filter.ref=f"A2:{get_column_letter(len(msg_cols))}2"
    rx_df=df[df["Direction"]=="Rx"].reset_index(drop=True)
    for ri,(_, row) in enumerate(rx_df.iterrows(), start=3):
        bg="0C1E35" if ri%2==0 else "0f1e33"
        for ci,cname in enumerate([c[0] for c in msg_cols],1):
            mapped={"Dir":"Direction","Ch":"Channel"}.get(cname,cname)
            val=row.get(mapped,"")
            if val is None: val=""
            if cname=="Dir": val="Rx"
            cell=ws6.cell(ri,ci,val)
            cell.font=bf(9,color=C_RX); cell.fill=fl(bg)
            cell.alignment=al() if cname in ("Timestamp","Data (Hex)","Rel Time (ms)") else ac()
            cell.border=brd

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-logo">🚌 CAN Analyzer</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Upload BUSMASTER Log",
        type=["log","txt"],
        label_visibility="collapsed",
        help="BUSMASTER .log / .txt format"
    )

    st.markdown('<div class="sec-header">⚙️ Filters</div>', unsafe_allow_html=True)
    filter_placeholder = st.empty()
    st.markdown("---")
    st.markdown('<div class="sec-header">ℹ️ About</div>', unsafe_allow_html=True)
    st.markdown("""
<div style="font-size:0.75rem;color:#8b949e;line-height:1.6;">
Parses <b>BUSMASTER</b> CAN logs.<br>
Supports Standard, Extended (29-bit), and Remote frames.<br><br>
Excel export includes:<br>
• Summary sheet<br>
• All messages + filters<br>
• <b>Timing / frequency analysis</b><br>
• ID-level breakdown<br>
• Tx / Rx split sheets
</div>
""", unsafe_allow_html=True)


# ─── MAIN ──────────────────────────────────────────────────────────────────────
if not uploaded:
    st.markdown("""
<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
            min-height:70vh;text-align:center;">
  <div style="font-size:3rem;margin-bottom:1rem;">🚌</div>
  <div style="font-size:1.5rem;font-weight:700;color:#e6edf3;margin-bottom:0.5rem;">
    CAN Bus Log Analyzer
  </div>
  <div style="font-size:0.9rem;color:#8b949e;max-width:420px;line-height:1.6;">
    Upload a <b style="color:#58a6ff;">BUSMASTER .log</b> file using the sidebar.<br>
    Get message inspection, timing analysis, payload decoding,<br>and a fully formatted Excel report.
  </div>
  <div style="margin-top:2rem;display:flex;gap:2rem;font-size:0.8rem;color:#8b949e;">
    <span>📋 Message Browser</span>
    <span>⏱️ Timing Analysis</span>
    <span>📊 ID Breakdown</span>
    <span>📥 Excel Export</span>
  </div>
</div>
""", unsafe_allow_html=True)
    st.stop()


# ── Parse ──
raw = uploaded.read().decode("utf-8", errors="replace")
with st.spinner("Parsing…"):
    meta, df = parse_can_log(raw)

if df.empty:
    st.error("No valid CAN messages found.")
    st.stop()

# ── Sidebar Filters (now that we have data) ──
with filter_placeholder.container():
    dir_opts = sorted(df["Direction"].unique().tolist())
    id_opts  = sorted(df["CAN ID"].unique().tolist())
    ft_opts  = sorted(df["Frame"].unique().tolist())

    sel_dir = st.multiselect("Direction", dir_opts, default=dir_opts)
    sel_id  = st.multiselect("CAN ID",    id_opts,  default=id_opts)
    sel_ft  = st.multiselect("Frame Type",ft_opts,  default=ft_opts)

    ts_range = None
    if df["Rel Time (ms)"].notna().any():
        t_min = float(df["Rel Time (ms)"].min())
        t_max = float(df["Rel Time (ms)"].max())
        ts_range = st.slider(
            "Time Window (ms)",
            min_value=t_min, max_value=t_max,
            value=(t_min, t_max), step=10.0,
        )

filt = (
    df["Direction"].isin(sel_dir) &
    df["CAN ID"].isin(sel_id) &
    df["Frame"].isin(sel_ft)
)
if ts_range:
    filt &= (df["Rel Time (ms)"] >= ts_range[0]) & (df["Rel Time (ms)"] <= ts_range[1])

fdf = df[filt].copy()


# ── KPI Row ──
dur_ms = df["ts_ms"].max() - df["ts_ms"].min() if df["ts_ms"].notna().any() else 0
dur_s  = dur_ms / 1000

st.markdown(f"""
<div class="kpi-row">
  <div class="kpi-card total">
    <div class="kpi-val">{len(df):,}</div>
    <div class="kpi-label">Total Messages</div>
  </div>
  <div class="kpi-card tx">
    <div class="kpi-val">{len(df[df["Direction"]=="Tx"]):,}</div>
    <div class="kpi-label">Transmitted (Tx)</div>
  </div>
  <div class="kpi-card rx">
    <div class="kpi-val">{len(df[df["Direction"]=="Rx"]):,}</div>
    <div class="kpi-label">Received (Rx)</div>
  </div>
  <div class="kpi-card ids">
    <div class="kpi-val">{df["CAN ID"].nunique()}</div>
    <div class="kpi-label">Unique CAN IDs</div>
  </div>
  <div class="kpi-card ext">
    <div class="kpi-val">{len(df[df["Frame"]=="Extended"])}</div>
    <div class="kpi-label">Extended Frames</div>
  </div>
  <div class="kpi-card dur">
    <div class="kpi-val">{round(dur_s,1)}s</div>
    <div class="kpi-label">Log Duration</div>
  </div>
</div>
""", unsafe_allow_html=True)

# Active filter notice
if len(fdf) < len(df):
    st.markdown(f'<div class="warn-banner">⚠️ Filter active — showing <b>{len(fdf):,}</b> of <b>{len(df):,}</b> messages</div>', unsafe_allow_html=True)


# ── Tabs ──
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📋  Messages", "⏱️  Timing", "🔍  ID Breakdown", "📡  Session Info", "📥  Export"
])


# ═══════ TAB 1 — MESSAGES ════════════════════════════════════════════════════
with tab1:
    c1, c2 = st.columns([3, 1])
    with c1:
        st.markdown(f'<div class="sec-header">CAN Messages &nbsp;<span style="color:#58a6ff;font-weight:700;">{len(fdf):,}</span> rows</div>', unsafe_allow_html=True)
    with c2:
        search = st.text_input("🔎 Search CAN ID / Data", placeholder="e.g. 0x302 or 05 00", label_visibility="collapsed")

    if search.strip():
        mask = (
            fdf["CAN ID"].str.contains(search.strip(), case=False, na=False) |
            fdf["Data (Hex)"].str.contains(search.strip(), case=False, na=False)
        )
        fdf = fdf[mask]
        st.caption(f"{len(fdf)} results for `{search}`")

    display_cols = ["Timestamp", "Rel Time (ms)", "Direction", "Channel", "CAN ID", "Frame", "DLC", "Data (Hex)",
                    "B0","B1","B2","B3","B4","B5","B6","B7"]

    def color_rows(row):
        if row["Direction"] == "Tx":
            return ["background-color:#0d2818; color:#3fb950"]*len(row)
        elif row["Direction"] == "Rx":
            return ["background-color:#0c1e35; color:#58a6ff"]*len(row)
        return [""]*len(row)

    def color_frame(val):
        if val == "Extended": return "color:#d2a8ff; font-weight:600"
        if val == "Remote":   return "color:#e3b341; font-weight:600"
        return "color:#c3c3e0"

    styled = (
        fdf[display_cols]
        .style
        .apply(color_rows, axis=1)
        .map(color_frame, subset=["Frame"])
        .format({"Rel Time (ms)": "{:.1f}"}, na_rep="—")
    )
    st.dataframe(styled, use_container_width=True, height=460, hide_index=True)

    # Payload inspector
    st.markdown('<div class="sec-header" style="margin-top:1.2rem;">🔬 Payload Inspector — Non-Zero Bytes</div>', unsafe_allow_html=True)
    byte_cols = ["B0","B1","B2","B3","B4","B5","B6","B7"]
    non_zero_df = fdf[fdf[byte_cols].apply(
        lambda r: any(v not in ("00","--","") for v in r), axis=1
    )][["Timestamp","Rel Time (ms)","Direction","CAN ID","Frame","DLC"] + byte_cols]

    if non_zero_df.empty:
        st.markdown('<div class="info-box">ℹ️ All payloads in current filter are <b>0x00</b>. This is typical for idle/heartbeat messages.</div>', unsafe_allow_html=True)
    else:
        st.caption(f"{len(non_zero_df)} messages with non-zero bytes")
        st.dataframe(non_zero_df.style.apply(color_rows, axis=1), use_container_width=True, height=280, hide_index=True)


# ═══════ TAB 2 — TIMING ══════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="sec-header">⏱️ Per-ID Timing & Frequency Analysis</div>', unsafe_allow_html=True)

    ta = timing_analysis(df)

    if ta.empty:
        st.warning("Not enough data for timing analysis.")
    else:
        # Visual timing bars
        max_mean = ta["Mean (ms)"].max()
        for _, row in ta.iterrows():
            pct = int(row["Mean (ms)"] / max_mean * 100) if max_mean > 0 else 0
            d_color = "#3fb950" if row["Direction"]=="Tx" else "#58a6ff"
            ft_badge = {"Extended":"ext","Remote":"rem","Standard":"std"}.get(row["Frame"],"std")
            st.markdown(f"""
<div class="timing-row">
  <div class="timing-id" style="color:{d_color};">{row["CAN ID"]}</div>
  <span class="badge-{ft_badge.lower()}">{row["Frame"]}</span>
  <div class="timing-bar-wrap">
    <div class="timing-bar" style="width:{pct}%;background:{d_color};opacity:0.7;"></div>
  </div>
  <div class="timing-val">
    <b style="color:{d_color};">{row["Est. Freq"]}</b>
    &nbsp;&nbsp;~{row["Mean (ms)"]} ms avg
  </div>
  <div style="font-size:0.72rem;color:#8b949e;min-width:80px;">jitter: {row["Jitter (ms)"]} ms</div>
  <div style="font-size:0.72rem;color:#8b949e;min-width:60px;">{row["Count"]} msgs</div>
</div>
""", unsafe_allow_html=True)

        st.markdown('<div class="sec-header" style="margin-top:1.5rem;">📊 Detailed Timing Table</div>', unsafe_allow_html=True)

        def _color_timing(val, max_val, mode="mean"):
            if max_val == 0 or not isinstance(val, (int, float)): return ""
            ratio = min(val / max_val, 1.0)
            if mode == "mean":
                r = int(30 + ratio*200); g = int(180 - ratio*130); b = 50
            else:
                r = int(30 + ratio*220); g = int(60 - ratio*40);   b = 50
            return f"background-color:rgb({r},{g},{b});color:#e6edf3;"

        _mm = ta["Mean (ms)"].max()
        _jm = ta["Jitter (ms)"].max()

        def _style_ta(row):
            out = [""] * len(row)
            for i, c in enumerate(row.index):
                if c == "Mean (ms)":   out[i] = _color_timing(row[c], _mm, "mean")
                elif c == "Jitter (ms)": out[i] = _color_timing(row[c], _jm, "jitter")
            return out

        st.dataframe(
            ta.style.apply(_style_ta, axis=1)
                    .format({"Mean (ms)":"{:.2f}","Min (ms)":"{:.2f}","Max (ms)":"{:.2f}",
                             "Std Dev (ms)":"{:.2f}","Jitter (ms)":"{:.2f}"}),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown('<div class="sec-header" style="margin-top:1rem;">📈 Message Rate Over Time</div>', unsafe_allow_html=True)
        if df["Rel Time (ms)"].notna().any():
            bucket_df = df.copy()
            bucket_df["sec"] = (bucket_df["Rel Time (ms)"] // 500).astype(int)  # 500ms buckets
            rate_df = bucket_df.groupby(["sec","CAN ID"]).size().unstack(fill_value=0)
            rate_df.index = rate_df.index * 0.5  # convert to seconds
            rate_df.index.name = "Time (s)"
            st.line_chart(rate_df, use_container_width=True, height=220)


# ═══════ TAB 3 — ID BREAKDOWN ════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="sec-header">🔍 CAN ID Traffic Breakdown</div>', unsafe_allow_html=True)

    total_msgs = len(df)
    for cid, grp in df.groupby("CAN ID"):
        pct = round(len(grp)/total_msgs*100, 1)
        tx_n = len(grp[grp["Direction"]=="Tx"])
        rx_n = len(grp[grp["Direction"]=="Rx"])
        ft   = grp["Frame"].mode()[0] if not grp.empty else "-"
        unique_pl = grp["Data (Hex)"].nunique()
        ft_badge = {"Extended":"ext","Remote":"rem","Standard":"std"}.get(ft,"std")
        d_color  = "#3fb950" if tx_n > rx_n else "#58a6ff"

        with st.expander(f"**{cid}** — {len(grp):,} msgs ({pct}%)  |  {ft}  |  Tx:{tx_n}  Rx:{rx_n}"):
            ic1, ic2, ic3, ic4 = st.columns(4)
            ic1.metric("Total",  f"{len(grp):,}")
            ic2.metric("Tx",     f"{tx_n:,}",  delta=f"{round(tx_n/len(grp)*100)}%")
            ic3.metric("Rx",     f"{rx_n:,}",  delta=f"{round(rx_n/len(grp)*100)}%")
            ic4.metric("Unique Payloads", unique_pl)

            # Timing for this ID
            sub = grp.sort_values("ts_ms")
            diffs = sub["ts_ms"].diff().dropna()
            if len(diffs) > 1:
                st.markdown(f"""
<div style="display:flex;gap:12px;flex-wrap:wrap;font-size:0.8rem;margin:8px 0;">
  <span style="color:#8b949e;">Mean interval: <b style="color:#e6edf3;">{diffs.mean():.1f} ms</b></span>
  <span style="color:#8b949e;">Freq: <b style="color:#3fb950;">{1000/diffs.mean():.1f} Hz</b></span>
  <span style="color:#8b949e;">Jitter: <b style="color:#e3b341;">{diffs.std():.1f} ms</b></span>
  <span style="color:#8b949e;">Min: <b style="color:#e6edf3;">{diffs.min():.0f} ms</b></span>
  <span style="color:#8b949e;">Max: <b style="color:#e6edf3;">{diffs.max():.0f} ms</b></span>
</div>
""", unsafe_allow_html=True)

            # Payload preview — byte-level
            st.markdown("**Payload samples:**")
            sample = grp[["Timestamp","Direction","Data (Hex)","B0","B1","B2","B3","B4","B5","B6","B7"]].head(8)
            st.dataframe(sample, use_container_width=True, hide_index=True, height=200)


# ═══════ TAB 4 — SESSION INFO ════════════════════════════════════════════════
with tab4:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="sec-header">📡 Session Metadata</div>', unsafe_allow_html=True)
        rows_html = "".join(f"<tr><td>{k}</td><td><b>{v}</b></td></tr>" for k,v in meta.items())
        st.markdown(f'<table class="meta-table">{rows_html}</table>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="sec-header">📊 Frame Type Distribution</div>', unsafe_allow_html=True)
        ft_counts = df["Frame"].value_counts().reset_index()
        ft_counts.columns = ["Frame Type", "Count"]
        ft_counts["% Share"] = (ft_counts["Count"]/len(df)*100).round(1).astype(str) + "%"
        st.dataframe(ft_counts, hide_index=True, use_container_width=True)

        st.markdown('<div class="sec-header" style="margin-top:1rem;">📊 Tx / Rx by CAN ID</div>', unsafe_allow_html=True)
        txrx = df.groupby(["CAN ID","Direction"]).size().unstack(fill_value=0)
        st.bar_chart(txrx, use_container_width=True, height=200)

    st.markdown('<div class="sec-header" style="margin-top:1rem;">🗂️ All IDs Summary</div>', unsafe_allow_html=True)
    summary = df.groupby("CAN ID").agg(
        Count=("CAN ID","count"),
        Tx=("Direction", lambda x:(x=="Tx").sum()),
        Rx=("Direction", lambda x:(x=="Rx").sum()),
        Frame=("Frame", lambda x:x.mode()[0]),
        Min_DLC=("DLC","min"), Max_DLC=("DLC","max"),
        First=("Timestamp","first"), Last=("Timestamp","last"),
        Unique_Payloads=("Data (Hex)","nunique")
    ).reset_index()
    summary["% Traffic"] = (summary["Count"]/len(df)*100).round(1).astype(str)+"%"
    st.dataframe(summary, hide_index=True, use_container_width=True, height=280)


# ═══════ TAB 5 — EXPORT ══════════════════════════════════════════════════════
with tab5:
    st.markdown('<div class="sec-header">📥 Download Excel Report</div>', unsafe_allow_html=True)

    st.markdown("""
<div class="info-box">
The Excel file is built with <b>Consolas monospace font</b>, dark-mode colour coding, freeze-panes on all sheets, and auto-filters. It contains 6 sheets:
<br><br>
<b>1. Summary</b> — session metadata + log statistics<br>
<b>2. All Messages</b> — every frame with relative timestamp + all 8 byte columns<br>
<b>3. Timing Analysis</b> — per-ID mean/min/max/jitter + estimated frequency (colour-scaled)<br>
<b>4. ID Breakdown</b> — traffic share, unique payloads, first/last seen<br>
<b>5. Tx Messages</b> — transmitted frames only<br>
<b>6. Rx Messages</b> — received frames only
</div>
""", unsafe_allow_html=True)

    with st.spinner("Building Excel…"):
        excel_bytes = build_excel(meta, df)

    fname = uploaded.name.replace(".log","").replace(".txt","")
    st.download_button(
        label="⬇️  Download  " + fname + "_CAN_Analysis.xlsx",
        data=excel_bytes,
        file_name=fname + "_CAN_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    st.markdown('<div class="sec-header" style="margin-top:1.5rem;">📋 Export Preview — Timing Analysis</div>', unsafe_allow_html=True)
    st.dataframe(timing_analysis(df), hide_index=True, use_container_width=True)